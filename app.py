from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, make_response, jsonify
from flask_sqlalchemy import SQLAlchemy
from datetime import date, datetime
from collections import defaultdict
from io import BytesIO
from xhtml2pdf import pisa
from openpyxl import Workbook
from sqlalchemy.orm import joinedload
import os
from werkzeug.utils import secure_filename
from datetime import date
import csv
from io import StringIO
import json


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///eglise.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = "change-me-en-cle-secrete"  # à changer pour la prod

UPLOAD_FOLDER_MEMBRES = os.path.join('static', 'uploads', 'membres')
os.makedirs(UPLOAD_FOLDER_MEMBRES, exist_ok=True)

app.config['UPLOAD_FOLDER_MEMBRES'] = UPLOAD_FOLDER_MEMBRES
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5 Mo par sécurité


db = SQLAlchemy(app)

@app.context_processor
def utility_processor():
    def url_for_other_page(page):
        """
        Construit l'URL pour une autre page de pagination en conservant :
        - les paramètres de chemin (ex: session_id)
        - les paramètres GET (filtres)
        """
        args = {}
        if request.view_args:
            args.update(request.view_args)          # ex: session_id
        args.update(request.args.to_dict())         # filtres actuels
        args['page'] = page
        return url_for(request.endpoint, **args)

    return dict(url_for_other_page=url_for_other_page)




def parse_date(value):
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()
    
def generer_matricule_par_civilite(civilite: str) -> str:
    civilite = (civilite or "").strip().lower()

    # Définir le préfixe selon la civilité
    if civilite.startswith("mr"):
        prefix = "L."
    elif civilite.startswith("mme") or civilite == "mrme":
        prefix = "V."
    else:
        # fallback générique si une civilité inattendue arrive
        prefix = "X."

    # Récupérer le matricule maximum pour ce préfixe
    # Format attendu : L.001  /  V.123  / X.005
    like_expr = f"{prefix}%"

    dernier = (
        Membre.query
        .filter(Membre.matricule.like(like_expr))
        .order_by(Membre.matricule.desc())
        .first()
    )

    if dernier:
        try:
            # Extraire la partie numérique : "L.123" → "123"
            num = int(dernier.matricule.split(".")[1])
            new_num = num + 1
        except Exception:
            new_num = 1
    else:
        new_num = 1

    # Format : L.001 / V.045
    return f"{prefix}{new_num:04d}"


def _build_membres_filtered_query_for_export():
    """
    Construit la même requête que dans liste_membres,
    mais sans pagination, pour les exports XLSX / PDF.
    """

    # session_id peut être un id numérique ou la valeur spéciale "__NOSESSION__"
    raw_session = request.args.get('session_id', default='', type=str)
    no_session_filter = (raw_session == '__NOSESSION__')
    selected_session_id = None

    if not no_session_filter and raw_session:
        try:
            selected_session_id = int(raw_session)
        except ValueError:
            selected_session_id = None

    nom = request.args.get('nom', type=str)
    civilite = request.args.get('civilite', type=str)
    benediction_id = request.args.get('benediction_id', type=int)
    talent_id = request.args.get('talent_id', type=int)
    ministere_id = request.args.get('ministere_id', type=int)
    faritra_id = request.args.get('faritra_id', type=int)
    telephone = request.args.get('telephone', type=str)
    adresse = request.args.get('adresse', type=str)
    famille_id = request.args.get('famille_id', type=int)
    responsable_id = request.args.get('responsable_id', type=str)
    zoky_only = request.args.get('zoky', type=int)  # si tu as ce filtre

    # ============ CAS "SANS SESSION" ============
    if no_session_filter:
        # Membres qui n'ont AUCUNE inscription
        query = db.session.query(Membre).distinct()
        query = query.outerjoin(InscriptionSession, InscriptionSession.membre_id == Membre.matricule)
        query = query.filter(InscriptionSession.id.is_(None))

        # Filtres basés uniquement sur Membre / Famille / Bénédictions / Talents
        if nom:
            query = query.filter(Membre.nom.ilike(f"%{nom}%"))
        if civilite:
            query = query.filter(Membre.civilite.ilike(f"%{civilite}%"))
        if famille_id:
            query = query.filter(Membre.famille_id == famille_id)

        if benediction_id:
            query = query.join(MembreBenediction, MembreBenediction.membre_id == Membre.matricule)
            query = query.filter(MembreBenediction.benediction_id == benediction_id)

        if talent_id:
            query = query.join(MembreTalent, MembreTalent.membre_id == Membre.matricule)
            query = query.filter(MembreTalent.talent_id == talent_id)

        # Zoky olona (si activé)
        if zoky_only:
            from datetime import date as _date
            today = _date.today()
            limite = _date(today.year - 65, today.month, today.day)
            query = query.filter(Membre.date_naissance <= limite)

    # ============ CAS NORMAL (AVEC SESSION) ============
    else:
        query = db.session.query(Membre).distinct()
        query = query.join(InscriptionSession, InscriptionSession.membre_id == Membre.matricule)
        query = query.join(Session, Session.id == InscriptionSession.session_id)

        if selected_session_id:
            query = query.filter(Session.id == selected_session_id)
        else:
            query = query.filter(Session.est_cloture == False)

        if nom:
            query = query.filter(Membre.nom.ilike(f"%{nom}%"))
        if civilite:
            query = query.filter(Membre.civilite.ilike(f"%{civilite}%"))

        if adresse:
            query = query.filter(InscriptionSession.adresse.ilike(f"%{adresse}%"))
        if faritra_id:
            query = query.filter(InscriptionSession.faritra_id == faritra_id)
        if telephone:
            query = query.filter(InscriptionSession.telephone.ilike(f"%{telephone}%"))

        if benediction_id:
            query = query.join(MembreBenediction, MembreBenediction.membre_id == Membre.matricule)
            query = query.filter(MembreBenediction.benediction_id == benediction_id)

        if talent_id:
            query = query.join(MembreTalent, MembreTalent.membre_id == Membre.matricule)
            query = query.filter(MembreTalent.talent_id == talent_id)

        if ministere_id:
            query = query.join(SessionMinistereMembre, SessionMinistereMembre.membre_id == Membre.matricule)
            query = query.join(SessionMinistere, SessionMinistere.id == SessionMinistereMembre.session_ministere_id)
            query = query.join(Ministere, Ministere.id == SessionMinistere.ministere_id)
            query = query.filter(Ministere.id == ministere_id)

        if famille_id:
            query = query.filter(Membre.famille_id == famille_id)

        # Responsable
        if responsable_id == '__NONE__':
            query = query.filter(InscriptionSession.responsable_id.is_(None))
        elif responsable_id:
            query = query.filter(InscriptionSession.responsable_id == responsable_id)

        # Zoky olona (65+)
        if zoky_only:
            from datetime import date as _date
            today = _date.today()
            limite = _date(today.year - 65, today.month, today.day)
            query = query.filter(Membre.date_naissance <= limite)

    return query.order_by(Membre.nom)

    
def delete_all_membres_and_ministere_links():
    """
    Supprime TOUTES les données membres :
      - MembreBenediction
      - MembreTalent
      - InscriptionSession
      - SessionMinistereMembre
      - Membre
    """
    try:
        # 1) Supprimer les bénédictions associées aux membres
        MembreBenediction.query.delete(synchronize_session=False)

        # 2) Supprimer les talents associés aux membres
        MembreTalent.query.delete(synchronize_session=False)

        # 3) Supprimer les inscriptions aux sessions
        InscriptionSession.query.delete(synchronize_session=False)

        # 4) Supprimer les membres des ministères
        SessionMinistereMembre.query.delete(synchronize_session=False)

        # 5) Supprimer tous les membres
        Membre.query.delete(synchronize_session=False)

        db.session.commit()

    except Exception as e:
        db.session.rollback()
        raise RuntimeError(f"Erreur lors de la suppression des membres : {e}")

# =============== MODELES =============== #

class Famille(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)

    membres = db.relationship('Membre', back_populates='famille')


class Faritra(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)


class Benediction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)

    membres = db.relationship('MembreBenediction', back_populates='benediction')


class Talent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)

    membres = db.relationship('MembreTalent', back_populates='talent')


class Session(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date_debut = db.Column(db.Date, nullable=False)
    date_fin = db.Column(db.Date, nullable=False)
    active = db.Column(db.Boolean, default=True)
    est_cloture = db.Column(db.Boolean, default=False)

    inscriptions = db.relationship('InscriptionSession', back_populates='session')
    ministeres = db.relationship('SessionMinistere', back_populates='session')


class Ministere(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)

    sessions = db.relationship('SessionMinistere', back_populates='ministere')


class Role(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nom = db.Column(db.String(100), nullable=False)


class Membre(db.Model):
    __table_args__ = (
        db.UniqueConstraint('civilite', 'nom', 'date_naissance', name='uq_membre_identite'),
    )
    
    matricule = db.Column(db.String(20), primary_key=True)
    nom = db.Column(db.String(150), nullable=False)

    famille_id = db.Column(db.Integer, db.ForeignKey('famille.id'))
    famille = db.relationship('Famille', back_populates='membres')

    adresse_initiale = db.Column(db.String(255))
    civilite = db.Column(db.String(20))
    date_naissance = db.Column(db.Date)

    diakona_mpiahy = db.Column(db.String(150))
    type_membre = db.Column(db.String(50))

    confirmation = db.Column(db.Boolean, default=False)
    bapteme = db.Column(db.Boolean, default=False)

    carte = db.Column(db.Boolean, default=False)
    souhait_avoir_carte = db.Column(db.Boolean, default=False)

    date_inscription = db.Column(db.Date, default=date.today)
    date_depart = db.Column(db.Date)

    est_mort = db.Column(db.Boolean, default=False)
    date_mort = db.Column(db.Date)

    date_mise_a_jour = db.Column(db.Date, default=date.today)
    
    photo = db.Column(db.String(255), nullable=True)       # nom de fichier dans /static/uploads/membres
    commentaire = db.Column(db.Text, nullable=True)


    benedictions = db.relationship('MembreBenediction', back_populates='membre')
    talents = db.relationship('MembreTalent', back_populates='membre')
    inscriptions = db.relationship(
        'InscriptionSession',
        back_populates='membre',
        foreign_keys='InscriptionSession.membre_id'
    )
    ministeres = db.relationship('SessionMinistereMembre', back_populates='membre')

    @property
    def benedictions_noms(self):
        return ", ".join(
            mb.benediction.nom
            for mb in self.benedictions
            if mb.benediction and mb.benediction.nom
        )

    @property
    def talents_noms(self):
        return ", ".join(
            mt.talent.nom
            for mt in self.talents
            if mt.talent and mt.talent.nom
        )
        
    @property
    def age(self):
        """Retourne l'âge en années ou None si date_naissance absente."""
        if not self.date_naissance:
            return None
        today = date.today()
        years = today.year - self.date_naissance.year
        # Ajustement si l’anniversaire n’est pas encore passé cette année
        if (today.month, today.day) < (self.date_naissance.month, self.date_naissance.day):
            years -= 1
        return years


class MembreBenediction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    membre_id = db.Column(db.String(20), db.ForeignKey('membre.matricule'))
    benediction_id = db.Column(db.Integer, db.ForeignKey('benediction.id'))

    lieu = db.Column(db.String(150))
    date_obtention = db.Column(db.Date)

    membre = db.relationship('Membre', back_populates='benedictions')
    benediction = db.relationship('Benediction', back_populates='membres')


class MembreTalent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    membre_id = db.Column(db.String(20), db.ForeignKey('membre.matricule'))
    talent_id = db.Column(db.Integer, db.ForeignKey('talent.id'))

    commentaire = db.Column(db.String(255))

    membre = db.relationship('Membre', back_populates='talents')
    talent = db.relationship('Talent', back_populates='membres')


class InscriptionSession(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    session_id = db.Column(db.Integer, db.ForeignKey('session.id'), nullable=False)
    membre_id = db.Column(db.String(20), db.ForeignKey('membre.matricule'), nullable=False)

    adresse = db.Column(db.String(255))
    faritra_id = db.Column(db.Integer, db.ForeignKey('faritra.id'))
    telephone = db.Column(db.String(30))

    responsable_id = db.Column(db.String(20), db.ForeignKey('membre.matricule'))

    session = db.relationship('Session', back_populates='inscriptions')
    membre = db.relationship('Membre', foreign_keys=[membre_id], back_populates='inscriptions')
    faritra = db.relationship('Faritra')
    responsable = db.relationship('Membre', foreign_keys=[responsable_id])


class SessionMinistere(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    session_id = db.Column(db.Integer, db.ForeignKey('session.id'), nullable=False)
    ministere_id = db.Column(db.Integer, db.ForeignKey('ministere.id'), nullable=False)

    session = db.relationship('Session', back_populates='ministeres')
    ministere = db.relationship('Ministere', back_populates='sessions')

    roles = db.relationship('SessionMinistereRole', back_populates='session_ministere')
    membres = db.relationship('SessionMinistereMembre', back_populates='session_ministere')


class SessionMinistereRole(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    session_ministere_id = db.Column(db.Integer, db.ForeignKey('session_ministere.id'), nullable=False)
    role_id = db.Column(db.Integer, db.ForeignKey('role.id'), nullable=False)
    max_membres = db.Column(db.Integer, nullable=True)  # si None => plusieurs possibles
    ordre = db.Column(db.Integer, nullable=True)

    session_ministere = db.relationship('SessionMinistere', back_populates='roles')
    role = db.relationship('Role')


class SessionMinistereMembre(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    session_ministere_id = db.Column(db.Integer, db.ForeignKey('session_ministere.id'), nullable=False)
    membre_id = db.Column(db.String(20), db.ForeignKey('membre.matricule'), nullable=False)
    role_id = db.Column(db.Integer, db.ForeignKey('role.id'), nullable=False)

    session_ministere = db.relationship('SessionMinistere', back_populates='membres')
    membre = db.relationship('Membre', back_populates='ministeres')
    role = db.relationship('Role')


# =============== SEED =============== #

def seed_data():
    created = []

    if not Role.query.filter_by(nom='membre').first():
        db.session.add(Role(nom='membre'))
        created.append("Rôle 'membre'")

    if not Ministere.query.filter_by(nom='FDL').first():
        db.session.add(Ministere(nom='FDL'))
        created.append("Ministère 'FDL'")

    if not Faritra.query.first():
        db.session.add(Faritra(nom='Centre'))
        db.session.add(Faritra(nom='Nord'))
        db.session.add(Faritra(nom='Sud'))
        created.append("Faritra de base")

    if not Benediction.query.first():
        db.session.add(Benediction(nom="Baptême"))
        db.session.add(Benediction(nom="Confirmation"))
        db.session.add(Benediction(nom="Mariage"))
        created.append("Bénédictions de base")

    if not Talent.query.first():
        db.session.add(Talent(nom="Chant"))
        db.session.add(Talent(nom="Musique"))
        db.session.add(Talent(nom="Enseignement"))
        created.append("Talents de base")

    if created:
        db.session.commit()
        print("Seed exécuté :", ", ".join(created))
    else:
        print("Seed : rien à créer.")


# =============== ROUTES =============== #

@app.route('/')
def index():
    # On réutilise la logique du dashboard
    total_membres = Membre.query.count()
    total_familles = Famille.query.count()
    total_sessions = Session.query.count()
    sessions_actives = Session.query.filter_by(active=True, est_cloture=False).count()

    membres_actifs = Membre.query.filter(
        Membre.est_mort == False,
        Membre.date_depart.is_(None)
    ).count()

    last_session = (
        Session.query
        .order_by(Session.date_debut.desc())
        .first()
    )

    inscriptions_last_session = 0
    if last_session:
        inscriptions_last_session = InscriptionSession.query.filter_by(
            session_id=last_session.id
        ).count()

    total_ministeres = Ministere.query.count()

    return render_template(
        'dashboard.html',
        total_membres=total_membres,
        membres_actifs=membres_actifs,
        total_familles=total_familles,
        total_sessions=total_sessions,
        sessions_actives=sessions_actives,
        last_session=last_session,
        inscriptions_last_session=inscriptions_last_session,
        total_ministeres=total_ministeres
    )


@app.route('/dashboard')
def dashboard():
    return redirect(url_for('index'))

@app.route('/admin/delete_all_membres', methods=['POST'])
def admin_delete_all_membres():
    try:
        delete_all_membres_and_ministere_links()
        flash("Tous les membres et les affectations ministérielles ont été supprimés.", "success")
    except RuntimeError as e:
        flash(str(e), "error")

    return redirect(url_for('sessions_liste'))


# ---- Membres ---- #
@app.route('/import_membres', methods=['GET', 'POST'])
def import_membres():
    session_active = Session.query.filter_by(active=True).first()
    if not session_active:
        flash("Aucune session active.", "error")
        return redirect(url_for("sessions_liste"))

    errors = []
    success_rows = []
    stats = None

    if request.method == "POST":
        fichier = request.files.get("fichier")
        is_simulation = bool(request.form.get("simulation"))

        if not fichier or not fichier.filename:
            flash("Veuillez sélectionner un fichier CSV.", "error")
            return render_template("import_membres.html",
                                   session_active=session_active,
                                   stats=None, errors=None, success_rows=None)

        # --- Lecture du fichier ---
        try:
            raw = fichier.read().decode("utf-8-sig")
        except:
            raw = fichier.read().decode("latin-1")

        # Supprimer BOM éventuel dans le contenu
        raw = raw.replace("\ufeff", "")

        # Détection du séparateur
        try:
            dialect = csv.Sniffer().sniff(raw[:1024], delimiters=";,\t")
        except:
            dialect = csv.excel

        reader = csv.DictReader(StringIO(raw), dialect=dialect)

        # --- NORMALISATION DES NOMS DES COLONNES ---
        clean_fields = []
        for fn in reader.fieldnames:
            fn_clean = fn.replace('\ufeff', '').strip().lower()
            clean_fields.append(fn_clean)
        reader.fieldnames = clean_fields

        # Caches
        familles_cache = {}

        nb_membres_crees = 0
        nb_membres_mis_a_jour = 0
        nb_inscriptions_creees = 0
        nb_benedictions_ajoutees = 0

        for idx, row in enumerate(reader, start=2):
            row_errors = []

            # --- Lecture colonnes normalisées ---
            matricule = (row.get("matricule") or "").strip()
            if not matricule:
                row_errors.append("Matricule manquant.")
                errors.append({"ligne": idx, "matricule": "", "message": "; ".join(row_errors), "row": row})
                continue

            nom = (row.get("nom") or "").strip()
            adresse = (row.get("adresse") or "").strip()
            telephone = (row.get("téléphone") or row.get("telephone") or "").strip()
            faritra_id_str = (row.get("faritra_id") or "").strip()
            benedictions_str = (row.get("benediction") or "").strip()

            date_naissance_str = (row.get("date_naissance") or "").strip()
            type_str = (row.get("type") or "").strip()
            civilite = (row.get("civilite") or "").strip()
            famille_code = (row.get("famille") or "").strip()
            responsable_mat = (row.get("responsable_id") or "").strip()
            ministere_str = (row.get("ministere") or "").strip().upper()

            # Faritra
            faritra_id = None
            if faritra_id_str:
                try:
                    faritra_id = int(faritra_id_str)
                except:
                    row_errors.append(f"Faritra_id invalide : {faritra_id_str}")

            if telephone == "0":
                telephone = None

            # Civilité automatique si absente
            if not civilite:
                if matricule.startswith("V"):
                    civilite = "Mme"
                elif matricule.startswith("L"):
                    civilite = "Mr"

            # --- Gestion Famille ---
            famille_id = None
            if famille_code:
                if famille_code in familles_cache:
                    famille_id = familles_cache[famille_code]
                else:
                    fam = Famille.query.filter(Famille.nom.ilike(famille_code)).first()
                    if not fam:
                        fam = Famille(nom=famille_code)
                        db.session.add(fam)
                        db.session.flush()
                    familles_cache[famille_code] = fam.id
                    famille_id = fam.id

            # --- Membre ---
            membre = Membre.query.get(matricule)
            if not membre:
                membre = Membre(
                    matricule=matricule,
                    nom=nom,
                    adresse_initiale=adresse,
                    civilite=civilite,
                    type_membre=type_str or None,
                    est_mort=False
                )
                if famille_id:
                    membre.famille_id = famille_id

                if date_naissance_str:
                    try:
                        membre.date_naissance = datetime.fromisoformat(date_naissance_str).date()
                    except:
                        row_errors.append(f"Date de naissance invalide : {date_naissance_str}")

                db.session.add(membre)
                nb_membres_crees += 1
                action_membre = "Créé"

            else:
                if nom:
                    membre.nom = nom
                if adresse:
                    membre.adresse_initiale = adresse
                if civilite:
                    membre.civilite = civilite
                if type_str:
                    membre.type_membre = type_str
                if famille_id:
                    membre.famille_id = famille_id

                if date_naissance_str:
                    try:
                        membre.date_naissance = datetime.fromisoformat(date_naissance_str).date()
                    except:
                        row_errors.append(f"Date de naissance invalide : {date_naissance_str}")

                nb_membres_mis_a_jour += 1
                action_membre = "Mis à jour"

            db.session.flush()

            # --- Inscription session ---
            ins = InscriptionSession.query.filter_by(
                session_id=session_active.id,
                membre_id=membre.matricule
            ).first()

            if not ins:
                ins = InscriptionSession(
                    session_id=session_active.id,
                    membre_id=membre.matricule,
                    adresse=adresse,
                    faritra_id=faritra_id,
                    telephone=telephone,
                    responsable_id=responsable_mat or None
                )
                db.session.add(ins)
                nb_inscriptions_creees += 1
            else:
                if adresse:
                    ins.adresse = adresse
                if faritra_id is not None:
                    ins.faritra_id = faritra_id
                if telephone:
                    ins.telephone = telephone
                if responsable_mat:
                    ins.responsable_id = responsable_mat

            # --- Ministère FDL ---
            if ministere_str == "FDL":
                ministere_fdl = Ministere.query.filter(Ministere.nom.ilike("FDL")).first()
                if not ministere_fdl:
                    ministere_fdl = Ministere(nom="FDL")
                    db.session.add(ministere_fdl)
                    db.session.flush()

                sm_fdl = SessionMinistere.query.filter_by(
                    session_id=session_active.id,
                    ministere_id=ministere_fdl.id
                ).first()

                if not sm_fdl:
                    sm_fdl = SessionMinistere(
                        session_id=session_active.id,
                        ministere_id=ministere_fdl.id
                    )
                    db.session.add(sm_fdl)
                    db.session.flush()

                role_membre = Role.query.filter(Role.nom.ilike("membre")).first()
                if not role_membre:
                    role_membre = Role(nom="membre")
                    db.session.add(role_membre)
                    db.session.flush()

                smr = SessionMinistereRole.query.filter_by(
                    session_ministere_id=sm_fdl.id,
                    role_id=role_membre.id
                ).first()

                if not smr:
                    smr = SessionMinistereRole(
                        session_ministere_id=sm_fdl.id,
                        role_id=role_membre.id,
                        max_membres=None,
                        ordre=None
                    )
                    db.session.add(smr)
                    db.session.flush()

                deja = SessionMinistereMembre.query.filter_by(
                    session_ministere_id=sm_fdl.id,
                    membre_id=membre.matricule
                ).first()

                if not deja:
                    smm = SessionMinistereMembre(
                        session_ministere_id=sm_fdl.id,
                        membre_id=membre.matricule,
                        role_id=role_membre.id
                    )
                    db.session.add(smm)

            # --- Bénédictions ---
            if benedictions_str:
                noms_ben = [b.strip() for b in benedictions_str.split(",") if b.strip()]
                for ben_nom in noms_ben:
                    ben = Benediction.query.filter(Benediction.nom.ilike(ben_nom)).first()
                    if not ben:
                        ben = Benediction(nom=ben_nom)
                        db.session.add(ben)
                        db.session.flush()
                    lien = MembreBenediction.query.filter_by(
                        membre_id=membre.matricule,
                        benediction_id=ben.id
                    ).first()
                    if not lien:
                        lien = MembreBenediction(membre_id=membre.matricule,
                                                 benediction_id=ben.id)
                        db.session.add(lien)
                        nb_benedictions_ajoutees += 1

            # Enregistrer erreurs ou succès
            if row_errors:
                errors.append({"ligne": idx, "matricule": matricule,
                               "message": "; ".join(row_errors), "row": row})
            else:
                success_rows.append({"ligne": idx, "matricule": matricule,
                                     "nom": nom, "action": action_membre,
                                     "row": row})

        # --- Commit ou rollback en cas de simulation ---
        if is_simulation:
            db.session.rollback()
        else:
            db.session.commit()

        stats = {
            "simulation": is_simulation,
            "membres_crees": nb_membres_crees,
            "membres_mis_a_jour": nb_membres_mis_a_jour,
            "inscriptions_creees": nb_inscriptions_creees,
            "benedictions_ajoutees": nb_benedictions_ajoutees,
            "nb_erreurs": len(errors),
            "nb_success": len(success_rows)
        }

        session["import_membres_errors"] = json.dumps(errors, ensure_ascii=False)
        session["import_membres_success"] = json.dumps(success_rows, ensure_ascii=False)

        return render_template("import_membres.html",
                               session_active=session_active,
                               stats=stats,
                               errors=errors,
                               success_rows=success_rows)

    return render_template("import_membres.html",
                           session_active=session_active,
                           stats=None, errors=None, success_rows=None)


@app.route('/import_membres/export')
def import_membres_export():
    kind = request.args.get('type', 'errors')  # 'errors' | 'success' | 'all'

    errors_json = session.get('import_membres_errors')
    success_json = session.get('import_membres_success')

    if not errors_json and not success_json:
        flash("Aucune donnée d'import à exporter. Relancez un import.", "warning")
        return redirect(url_for('import_membres'))

    errors = json.loads(errors_json) if errors_json else []
    success_rows = json.loads(success_json) if success_json else []

    # Préparation des données selon le type demandé
    if kind == 'errors':
        data = [
            {
                "status": "ERREUR",
                "ligne": e.get("ligne"),
                "matricule": e.get("matricule"),
                "nom": (e.get("row") or {}).get("nom", ""),
                "action": "",
                "message": e.get("message"),
                "row": e.get("row") or {}
            }
            for e in errors
        ]
        filename = "import_membres_erreurs.csv"

    elif kind == 'success':
        data = [
            {
                "status": "OK",
                "ligne": s.get("ligne"),
                "matricule": s.get("matricule"),
                "nom": s.get("nom"),
                "action": s.get("action"),
                "message": "",
                "row": s.get("row") or {}
            }
            for s in success_rows
        ]
        filename = "import_membres_success.csv"

    else:  # kind == 'all' -> on fusionne les deux
        data = []
        for e in errors:
            data.append({
                "status": "ERREUR",
                "ligne": e.get("ligne"),
                "matricule": e.get("matricule"),
                "nom": (e.get("row") or {}).get("nom", ""),
                "action": "",
                "message": e.get("message"),
                "row": e.get("row") or {}
            })
        for s in success_rows:
            data.append({
                "status": "OK",
                "ligne": s.get("ligne"),
                "matricule": s.get("matricule"),
                "nom": s.get("nom"),
                "action": s.get("action"),
                "message": "",
                "row": s.get("row") or {}
            })
        filename = "import_membres_rapport_complet.csv"

    if not data:
        flash("Aucune donnée à exporter pour ce type.", "warning")
        return redirect(url_for('import_membres'))

    # Construire l'ensemble des colonnes originales
    all_row_keys = set()
    for item in data:
        row_data = item.get("row") or {}
        all_row_keys.update(row_data.keys())

    # On s'assure d'un ordre stable
    all_row_keys = sorted(all_row_keys)

    # En-tête CSV : meta + colonnes originales
    header = ['ligne', 'status', 'matricule', 'nom', 'action', 'message'] + list(all_row_keys)

    # Génération du CSV en mémoire
    output = StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(header)

    for item in data:
        base = [
            item.get("ligne", ""),
            item.get("status", ""),
            item.get("matricule", ""),
            item.get("nom", ""),
            item.get("action", ""),
            item.get("message", "")
        ]
        row_data = item.get("row") or {}
        for key in all_row_keys:
            base.append(row_data.get(key, ""))
        writer.writerow(base)

    csv_content = output.getvalue()
    output.close()

    response = make_response(csv_content)
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response

@app.route('/membres', methods=['GET', 'POST'])
def liste_membres():
    # --------- GESTION DES ACTIONS GROUPÉES (POST) ---------
    if request.method == 'POST':
        bulk_action = request.form.get('bulk_action', '').strip()
        selected_matricules = request.form.getlist('selected_membres')

        if not bulk_action:
            flash("Veuillez choisir une action groupée.", "error")
            return redirect(url_for('liste_membres', **request.args.to_dict()))

        if not selected_matricules:
            flash("Veuillez sélectionner au moins un membre.", "error")
            return redirect(url_for('liste_membres', **request.args.to_dict()))

        # On récupère la session courante pour les actions qui en ont besoin
        selected_session_id = request.args.get('session_id', type=int)
        if not selected_session_id:
            session_active = Session.query.filter_by(active=True, est_cloture=False).first()
            selected_session_id = session_active.id if session_active else None

        # ---- 1) Action : créer/associer une famille ----
        if bulk_action == 'creer_famille':
            famille_id = request.form.get('bulk_famille_id') or None
            nouvelle_famille_nom = (request.form.get('bulk_nouvelle_famille') or '').strip()

            if not famille_id and not nouvelle_famille_nom:
                flash("Pour créer une famille, choisissez une famille existante ou indiquez un nom de nouvelle famille.", "error")
                return redirect(url_for('liste_membres', **request.args.to_dict()))

            # Si nouvelle famille => on la crée
            if not famille_id and nouvelle_famille_nom:
                fam = Famille.query.filter(Famille.nom.ilike(nouvelle_famille_nom)).first()
                if not fam:
                    fam = Famille(nom=nouvelle_famille_nom)
                    db.session.add(fam)
                    db.session.flush()
                famille_id = fam.id

            nb_modifies = 0
            for mat in selected_matricules:
                m = Membre.query.get(mat)
                if not m:
                    continue
                m.famille_id = famille_id
                nb_modifies += 1

            db.session.commit()
            flash(f"{nb_modifies} membre(s) ont été associés à la famille.", "success")
            return redirect(url_for('liste_membres', **request.args.to_dict()))

        # ---- 2) Action : affecter un responsable (FDL) ----
        if bulk_action == 'set_responsable':
            responsable_id = (request.form.get('bulk_responsable_id') or '').strip()
            if not responsable_id:
                flash("Veuillez choisir un responsable FDL pour cette action.", "error")
                return redirect(url_for('liste_membres', **request.args.to_dict()))

            if not selected_session_id:
                flash("Impossible d'affecter un responsable : aucune session courante définie.", "error")
                return redirect(url_for('liste_membres', **request.args.to_dict()))

            nb_affectes = 0
            for mat in selected_matricules:
                insc = InscriptionSession.query.filter_by(
                    session_id=selected_session_id,
                    membre_id=mat
                ).first()
                if not insc:
                    # pas inscrit à cette session → on ignore silencieusement
                    continue
                insc.responsable_id = responsable_id
                nb_affectes += 1

            db.session.commit()
            flash(f"Responsable affecté pour {nb_affectes} inscription(s).", "success")
            return redirect(url_for('liste_membres', **request.args.to_dict()))

        # ---- 3) Action : supprimer les membres NON inscrits à une session ----
        if bulk_action == 'supprimer_non_inscrits':
            nb_supprimes = 0
            nb_ignores = 0
            membres_ignores = []

            for mat in selected_matricules:
                m = Membre.query.get(mat)
                if not m:
                    continue

                # Vérifier s'il a des inscriptions à une ou plusieurs sessions
                nb_insc = InscriptionSession.query.filter_by(membre_id=mat).count()
                if nb_insc > 0:
                    nb_ignores += 1
                    membres_ignores.append(f"{mat} - {m.nom}")
                    continue

                # Par sécurité, on nettoie aussi les liens secondaires :
                MembreBenediction.query.filter_by(membre_id=mat).delete(synchronize_session=False)
                MembreTalent.query.filter_by(membre_id=mat).delete(synchronize_session=False)
                SessionMinistereMembre.query.filter_by(membre_id=mat).delete(synchronize_session=False)

                # Si ce membre était éventuellement responsable d'autres inscriptions,
                # on enlève la référence à lui comme responsable :
                InscriptionSession.query.filter_by(responsable_id=mat).update(
                    {InscriptionSession.responsable_id: None},
                    synchronize_session=False
                )

                db.session.delete(m)
                nb_supprimes += 1

            db.session.commit()

            if nb_supprimes > 0:
                flash(f"{nb_supprimes} membre(s) non inscrit(s) à une session ont été supprimés.", "success")

            if nb_ignores > 0:
                txt = ", ".join(membres_ignores[:5])
                if nb_ignores > 5:
                    txt += " …"
                flash(
                    f"{nb_ignores} membre(s) n'ont pas été supprimés car ils sont inscrits à au moins une session : {txt}",
                    "warning"
                )

            return redirect(url_for('liste_membres', **request.args.to_dict()))

        # Si action inconnue
        flash("Action groupée inconnue.", "error")
        return redirect(url_for('liste_membres', **request.args.to_dict()))

        # --------- PARTIE LISTE (GET) ---------
    page = request.args.get('page', 1, type=int)
    per_page = 20  # tu peux ajuster

    # session_id peut être un id numérique ou la valeur spéciale "__NOSESSION__"
    raw_session = request.args.get('session_id', default='', type=str)
    no_session_filter = (raw_session == '__NOSESSION__')
    selected_session_id = None

    if not no_session_filter and raw_session:
        try:
            selected_session_id = int(raw_session)
        except ValueError:
            selected_session_id = None

    nom = request.args.get('nom', type=str)
    civilite = request.args.get('civilite', type=str)
    benediction_id = request.args.get('benediction_id', type=int)
    talent_id = request.args.get('talent_id', type=int)
    ministere_id = request.args.get('ministere_id', type=int)
    faritra_id = request.args.get('faritra_id', type=int)
    telephone = request.args.get('telephone', type=str)
    adresse = request.args.get('adresse', type=str)
    famille_id = request.args.get('famille_id', type=int)
    responsable_id = request.args.get('responsable_id', type=str)
    zoky_only = request.args.get('zoky', type=int)  # si tu l'utilises

    sessions_non_cloturees = (
        Session.query
        .filter_by(est_cloture=False)
        .order_by(Session.date_debut.desc())
        .all()
    )

    # Si aucune session n'est passée ET pas de filtre "sans session" → on prend la session active
    if not raw_session and not no_session_filter:
        session_active = Session.query.filter_by(active=True, est_cloture=False).first()
        if session_active:
            selected_session_id = session_active.id
            raw_session = str(session_active.id)

    # ============ CONSTRUCTION DE LA REQUÊTE ============

    if no_session_filter:
        # Cas "Sans session" : membres qui n'ont AUCUNE inscription
        query = db.session.query(Membre).distinct()
        # on utilise la relation Membre.inscriptions
        query = query.outerjoin(InscriptionSession, InscriptionSession.membre_id == Membre.matricule)
        query = query.filter(InscriptionSession.id.is_(None))

        # Filtres qui ne dépendent que de Membre + bénédictions + talents + famille
        if nom:
            query = query.filter(Membre.nom.ilike(f"%{nom}%"))
        if civilite:
            query = query.filter(Membre.civilite.ilike(f"%{civilite}%"))
        if famille_id:
            query = query.filter(Membre.famille_id == famille_id)

        if benediction_id:
            query = query.join(MembreBenediction, MembreBenediction.membre_id == Membre.matricule)
            query = query.filter(MembreBenediction.benediction_id == benediction_id)

        if talent_id:
            query = query.join(MembreTalent, MembreTalent.membre_id == Membre.matricule)
            query = query.filter(MembreTalent.talent_id == talent_id)

        # On ignore volontairement les filtres basés sur InscriptionSession / Session / Ministère
        # (adresse, faritra, téléphone, responsable, ministere, etc.) car par définition
        # ces membres n'ont pas de session.

    else:
        # Cas normal : filtrage par session (comme avant)
        query = db.session.query(Membre).distinct()
        query = query.join(InscriptionSession, InscriptionSession.membre_id == Membre.matricule)
        query = query.join(Session, Session.id == InscriptionSession.session_id)

        if selected_session_id:
            query = query.filter(Session.id == selected_session_id)
        else:
            query = query.filter(Session.est_cloture == False)

        if nom:
            query = query.filter(Membre.nom.ilike(f"%{nom}%"))
        if civilite:
            query = query.filter(Membre.civilite.ilike(f"%{civilite}%"))

        if adresse:
            query = query.filter(InscriptionSession.adresse.ilike(f"%{adresse}%"))
        if faritra_id:
            query = query.filter(InscriptionSession.faritra_id == faritra_id)
        if telephone:
            query = query.filter(InscriptionSession.telephone.ilike(f"%{telephone}%"))

        if benediction_id:
            query = query.join(MembreBenediction, MembreBenediction.membre_id == Membre.matricule)
            query = query.filter(MembreBenediction.benediction_id == benediction_id)

        if talent_id:
            query = query.join(MembreTalent, MembreTalent.membre_id == Membre.matricule)
            query = query.filter(MembreTalent.talent_id == talent_id)

        if ministere_id:
            query = query.join(SessionMinistereMembre, SessionMinistereMembre.membre_id == Membre.matricule)
            query = query.join(SessionMinistere, SessionMinistere.id == SessionMinistereMembre.session_ministere_id)
            query = query.join(Ministere, Ministere.id == SessionMinistere.ministere_id)
            query = query.filter(Ministere.id == ministere_id)

        if famille_id:
            query = query.filter(Membre.famille_id == famille_id)

        # Filtre responsable :
        if responsable_id == '__NONE__':
            query = query.filter(InscriptionSession.responsable_id.is_(None))
        elif responsable_id:
            query = query.filter(InscriptionSession.responsable_id == responsable_id)

        # Filtre Zoky olona (65+)
        if zoky_only:
            from datetime import date as _date
            today = _date.today()
            limite = _date(today.year - 65, today.month, today.day)
            query = query.filter(Membre.date_naissance <= limite)

    query = query.order_by(Membre.nom)

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    membres = pagination.items
    total_membres = pagination.total

    toutes_benedictions = Benediction.query.order_by(Benediction.nom).all()
    tous_talents = Talent.query.order_by(Talent.nom).all()
    tous_ministeres = Ministere.query.order_by(Ministere.nom).all()
    tous_faritras = Faritra.query.order_by(Faritra.nom).all()
    familles = Famille.query.order_by(Famille.nom).all()

    if selected_session_id:
        try:
            responsables = get_responsables_fdl(selected_session_id)
        except NameError:
            responsables = (
                db.session.query(Membre)
                .join(InscriptionSession, InscriptionSession.membre_id == Membre.matricule)
                .filter(
                    InscriptionSession.session_id == selected_session_id,
                    InscriptionSession.responsable_id.isnot(None)
                )
                .distinct()
                .order_by(Membre.nom)
                .all()
            )
    else:
        responsables = []

    return render_template(
        'membres_liste.html',
        membres=membres,
        sessions=sessions_non_cloturees,
        selected_session_id=selected_session_id,
        no_session_filter=no_session_filter,
        total_membres=total_membres,
        toutes_benedictions=toutes_benedictions,
        tous_talents=tous_talents,
        tous_ministeres=tous_ministeres,
        tous_faritras=tous_faritras,
        familles=familles,
        nom=nom or '',
        civilite=civilite or '',
        benediction_id=benediction_id,
        talent_id=talent_id,
        ministere_id=ministere_id,
        faritra_id=faritra_id,
        telephone=telephone or '',
        adresse=adresse or '',
        famille_id=famille_id,
        responsable_id=responsable_id,
        responsables=responsables,
        zoky=zoky_only,
        pagination=pagination
    )



@app.route('/api/next_matricule')
def api_next_matricule():
    civilite = request.args.get('civilite', '').strip()
    if not civilite:
        return jsonify({"matricule": ""})
    matricule = generer_matricule_par_civilite(civilite)
    return jsonify({"matricule": matricule})


@app.route('/membres/nouveau', methods=['GET', 'POST'])
def nouveau_membre():
    familles = Famille.query.all()
    benedictions = Benediction.query.order_by(Benediction.nom).all()
    talents = Talent.query.order_by(Talent.nom).all()

    if request.method == 'POST':
        matricule = request.form.get('matricule', '').strip()
        nom = request.form.get('nom', '').strip()
        famille_id = request.form.get('famille_id') or None
        nouvelle_famille_nom = request.form.get('nouvelle_famille', '').strip()
        adresse_initiale = request.form.get('adresse_initiale', '').strip()
        civilite = request.form.get('civilite', '').strip()
        type_membre = request.form.get('type_membre', '').strip()
        carte = 'carte' in request.form
        souhait_avoir_carte = 'souhait_avoir_carte' in request.form

        benediction_ids = request.form.getlist('benedictions')
        talent_ids = request.form.getlist('talents')

        errors = []

        # Matricule : soit fourni, soit généré selon civilité
        if not matricule and not civilite:
            errors.append("Civilité obligatoire si le matricule est généré automatiquement.")
        if not nom:
            errors.append("Le nom est obligatoire.")
        if not adresse_initiale:
            errors.append("L'adresse initiale est obligatoire.")
        if not civilite:
            errors.append("La civilité est obligatoire.")
        if not type_membre:
            errors.append("Le type (Père / Mère / Enfant) est obligatoire.")
        if not famille_id and not nouvelle_famille_nom:
            errors.append("La famille est obligatoire (sélectionner une famille ou en créer une nouvelle).")

        # Date de naissance
        date_naissance_val = None
        date_naissance_str = request.form.get("date_naissance", "").strip()
        if date_naissance_str:
            try:
                date_naissance_val = date.fromisoformat(date_naissance_str)
            except ValueError:
                errors.append("Date de naissance invalide.")
        else:
            errors.append("La date de naissance est obligatoire.")
        
                # Unicité : civilité + nom + date de naissance
        if civilite and nom and date_naissance_val:
            existing = (
                Membre.query
                .filter(
                    Membre.civilite == civilite,
                    Membre.nom.ilike(nom),
                    Membre.date_naissance == date_naissance_val
                )
                .first()
            )
            if existing:
                errors.append(
                    f"Un membre avec cette civilité, ce nom et cette date de naissance existe déjà "
                    f"(matricule : {existing.matricule})."
                )

        if errors:
            for e in errors:
                flash(e, "error")
            return render_template(
                'membres_nouveau.html',
                familles=familles,
                benedictions=benedictions,
                talents=talents
            )


        if errors:
            for e in errors:
                flash(e, "error")
            return render_template(
                'membres_nouveau.html',
                familles=familles,
                benedictions=benedictions,
                talents=talents
            )

        # Si matricule non fourni → génération automatique
        if not matricule:
            matricule = generer_matricule_par_civilite(civilite)

        # Création éventuelle d'une nouvelle famille
        if not famille_id and nouvelle_famille_nom:
            nouvelle_famille = Famille(nom=nouvelle_famille_nom)
            db.session.add(nouvelle_famille)
            db.session.flush()
            famille_id = nouvelle_famille.id

        membre = Membre(
            matricule=matricule,
            nom=nom,
            famille_id=famille_id,
            adresse_initiale=adresse_initiale,
            civilite=civilite,
            type_membre=type_membre,
            date_naissance=date_naissance_val,
            carte=carte,
            souhait_avoir_carte=souhait_avoir_carte,
        )
        db.session.add(membre)

        # Attacher les bénédictions sélectionnées (sans lieu/date ici)
        for bid in benediction_ids:
            if bid:
                db.session.add(MembreBenediction(
                    membre_id=matricule,
                    benediction_id=int(bid)
                ))

        # Attacher les talents sélectionnés
        for tid in talent_ids:
            if tid:
                db.session.add(MembreTalent(
                    membre_id=matricule,
                    talent_id=int(tid)
                ))

        db.session.commit()
        flash(f"Membre créé avec succès (matricule : {matricule}).", "success")
        return redirect(url_for('liste_membres'))

    return render_template(
        'membres_nouveau.html',
        familles=familles,
        benedictions=benedictions,
        talents=talents
    )


@app.route('/familles/<int:famille_id>/fiche')
def famille_fiche(famille_id):
    famille = Famille.query.get_or_404(famille_id)

    def ordre_type(m):
        t = (m.type_membre or "").strip().lower()
        if t == "père" or t == "pere":
            return 0
        if t == "mère" or t == "mere":
            return 1
        if t == "enfant":
            return 2
        return 3  # autres types à la fin

    # Tri : d'abord type (Père, Mère, Enfant, autres), puis nom
    membres = sorted(
        famille.membres,
        key=lambda m: (ordre_type(m), (m.nom or "").lower())
    )

    return render_template('famille_fiche.html', famille=famille, membres=membres)

@app.route('/membres/export/xlsx')
def membres_export_xlsx():
    query = _build_membres_filtered_query_for_export()
    membres = query.all()

    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Membres"

    headers = [
        "Matricule", "Nom", "Famille", "Civilité",
        "Adresse (session)", "Faritra (session)", "Téléphone (session)",
        "Bénédictions", "Talents", "Ministères (session)"
    ]
    ws.append(headers)

    selected_session_id = request.args.get('session_id', type=int)

    for m in membres:
        insc = None
        if selected_session_id:
            for i in m.inscriptions:
                if i.session_id == selected_session_id:
                    insc = i
                    break

        ministeres_str = ""
        if selected_session_id:
            mins = []
            for smm in m.ministeres:
                sm = smm.session_ministere
                if sm and sm.session_id == selected_session_id:
                    label = sm.ministere.nom if sm.ministere else ""
                    if smm.role:
                        label += f" - {smm.role.nom}"
                    mins.append(label)
            ministeres_str = "; ".join(mins)

        ws.append([
            m.matricule,
            m.nom,
            m.famille.nom if m.famille else "",
            m.civilite or "",
            insc.adresse if insc else "",
            insc.faritra.nom if insc and insc.faritra else "",
            insc.telephone if insc else "",
            m.benedictions_noms,
            m.talents_noms,
            ministeres_str
        ])

    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="membres_filtrés.xlsx"
    )


@app.route('/membres/export/pdf')
def membres_export_pdf():
    query = _build_membres_filtered_query_for_export()
    membres = query.all()

    selected_session_id = request.args.get('session_id', type=int)

    html = render_template(
        'membres_export_pdf.html',
        membres=membres,
        selected_session_id=selected_session_id
    )

    pdf_io = BytesIO()
    pisa.CreatePDF(html, dest=pdf_io)
    pdf_io.seek(0)

    return send_file(
        pdf_io,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="membres_filtrés.pdf"
    )
    
@app.route('/membres/<matricule>/fiche/pdf')
def membre_fiche_pdf(matricule):
    membre = Membre.query.get_or_404(matricule)

    # même préparation que dans membre_fiche
    derniere_inscription = None
    if membre.inscriptions:
        derniere_inscription = sorted(
            membre.inscriptions,
            key=lambda i: i.id,
            reverse=True
        )[0]

    def session_sort_key(ins):
        s = ins.session
        if s and s.date_debut:
            return s.date_debut
        return date.min

    inscriptions_tries = sorted(
        membre.inscriptions,
        key=session_sort_key,
        reverse=True
    )

    ministeres_par_session = defaultdict(list)
    for smm in membre.ministeres:
        sm = smm.session_ministere
        if not sm or not sm.session:
            continue
        sess = sm.session
        ministeres_par_session[sess.id].append({
            "ministere": sm.ministere.nom if sm.ministere else "",
            "role": smm.role.nom if smm.role else "",
        })

    html = render_template(
        'membre_fiche.html',
        membre=membre,
        derniere_inscription=derniere_inscription,
        inscriptions_tries=inscriptions_tries,
        ministeres_par_session=ministeres_par_session
    )

    pdf_io = BytesIO()
    pisa.CreatePDF(html, dest=pdf_io)
    pdf_io.seek(0)

    filename = f"fiche_membre_{membre.matricule}.pdf"
    return send_file(
        pdf_io,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )

@app.route('/membres/<matricule>/fiche/xlsx')
def membre_fiche_xlsx(matricule):
    membre = Membre.query.get_or_404(matricule)

    output = BytesIO()
    wb = Workbook()

    # Feuille info
    ws_info = wb.active
    ws_info.title = "Infos"

    ws_info.append(["Champ", "Valeur"])
    ws_info.append(["Matricule", membre.matricule])
    ws_info.append(["Nom", membre.nom])
    ws_info.append(["Civilité", membre.civilite or ""])
    ws_info.append(["Type", membre.type_membre or ""])
    ws_info.append(["Famille", membre.famille.nom if membre.famille else ""])
    ws_info.append(["Adresse initiale", membre.adresse_initiale or ""])
    ws_info.append(["Date inscription", membre.date_inscription or ""])
    ws_info.append(["Date départ", membre.date_depart or ""])
    ws_info.append(["Est mort", "Oui" if membre.est_mort else "Non"])
    ws_info.append(["Date mort", membre.date_mort or ""])
    ws_info.append(["Carte", "Oui" if membre.carte else "Non"])
    ws_info.append(["Souhaite une carte", "Oui" if membre.souhait_avoir_carte else "Non"])
    ws_info.append(["Dernière mise à jour", membre.date_mise_a_jour or ""])

    # Feuille bénédictions
    ws_b = wb.create_sheet("Bénédictions")
    ws_b.append(["Bénédiction", "Lieu", "Date"])
    for mb in membre.benedictions:
        ws_b.append([
            mb.benediction.nom if mb.benediction else "",
            mb.lieu or "",
            mb.date_obtention or ""
        ])

    # Feuille talents
    ws_t = wb.create_sheet("Talents")
    ws_t.append(["Talent", "Commentaire"])
    for mt in membre.talents:
        ws_t.append([
            mt.talent.nom if mt.talent else "",
            mt.commentaire or ""
        ])

    # Feuille sessions / ministères
    ws_s = wb.create_sheet("Sessions")
    ws_s.append([
        "Session ID", "Date début", "Date fin",
        "Adresse", "Faritra", "Téléphone",
        "Responsable", "Ministères & rôles"
    ])

    for ins in sorted(membre.inscriptions, key=lambda i: i.id):
        s = ins.session
        mins = []
        for smm in membre.ministeres:
            sm = smm.session_ministere
            if sm and sm.session_id == (s.id if s else None):
                label = sm.ministere.nom if sm.ministere else ""
                if smm.role:
                    label += f" - {smm.role.nom}"
                mins.append(label)

        ws_s.append([
            s.id if s else "",
            s.date_debut if s else "",
            s.date_fin if s else "",
            ins.adresse or "",
            ins.faritra.nom if ins.faritra else "",
            ins.telephone or "",
            ins.responsable.nom if ins.responsable else "",
            "; ".join(mins)
        ])

    wb.save(output)
    output.seek(0)
    filename = f"fiche_membre_{membre.matricule}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

@app.route('/membres/<matricule>/modifier', methods=['GET', 'POST'])
def membre_modifier(matricule):
    benedictions = Benediction.query.order_by(Benediction.nom).all()
    talents = Talent.query.order_by(Talent.nom).all()
    familles = Famille.query.order_by(Famille.nom).all()
    membre = Membre.query.get_or_404(matricule)

    if request.method == 'POST':
        # --- Infos de base ---
        membre.nom = request.form.get('nom', membre.nom)
        membre.adresse_initiale = request.form.get('adresse_initiale', membre.adresse_initiale)
        membre.civilite = request.form.get('civilite', membre.civilite)

        # Type (Père / Mère / Enfant)
        type_str = request.form.get('type_membre') or request.form.get('type')
        if type_str is not None:
            membre.type_membre = type_str.strip() or None

        # Famille : select existant OU création
        famille_id_str = request.form.get('famille_id', '').strip()
        famille_nom = request.form.get('nouvelle_famille', '').strip()

        if famille_id_str:
            try:
                membre.famille_id = int(famille_id_str)
            except ValueError:
                flash("Famille sélectionnée invalide.", "error")
        elif famille_nom:
            fam = Famille.query.filter(Famille.nom.ilike(famille_nom)).first()
            if not fam:
                fam = Famille(nom=famille_nom)
                db.session.add(fam)
                db.session.flush()
            membre.famille_id = fam.id
        # sinon on laisse la famille actuelle

        # Date de naissance
        date_naissance_str = request.form.get('date_naissance', '').strip()
        if date_naissance_str:
            try:
                membre.date_naissance = date.fromisoformat(date_naissance_str)
            except ValueError:
                flash("Date de naissance invalide.", "error")
        else:
            membre.date_naissance = None

        # Date départ
        date_depart_str = request.form.get('date_depart', '').strip()
        if date_depart_str:
            try:
                membre.date_depart = date.fromisoformat(date_depart_str)
            except ValueError:
                flash("Date de départ invalide.", "error")
        else:
            membre.date_depart = None

        # Commentaire
        membre.commentaire = request.form.get('commentaire', '').strip() or None

        # Est mort + date mort auto
        est_mort_bool = bool(request.form.get('est_mort'))
        membre.est_mort = est_mort_bool
        if membre.est_mort:
            if not membre.date_mort:
                membre.date_mort = date.today()
        else:
            membre.date_mort = None

        # Carte / souhait carte
        membre.carte = bool(request.form.get('carte'))
        membre.souhait_avoir_carte = bool(request.form.get('souhait_avoir_carte'))

        # Upload photo
        photo_file = request.files.get('photo')
        if photo_file and photo_file.filename:
            filename = secure_filename(f"{membre.matricule}_{photo_file.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER_MEMBRES'], filename)
            photo_file.save(filepath)
            membre.photo = filename

        # =====================================================
        # 1) GESTION DES BÉNÉDICTIONS (tableau mb_*)
        # =====================================================
        mb_ids = request.form.getlist('mb_id')                    # hidden (id de MembreBenediction ou vide)
        mb_benediction_ids = request.form.getlist('mb_benediction_id')
        mb_lieux = request.form.getlist('mb_lieu')
        mb_dates = request.form.getlist('mb_date')

        for idx in range(len(mb_ids)):
            mb_id_str = (mb_ids[idx] or "").strip()
            ben_id_str = (mb_benediction_ids[idx] or "").strip()
            lieu = (mb_lieux[idx] or "").strip()
            date_str = (mb_dates[idx] or "").strip()

            # Si tout est vide => ligne purement vide => on ignore
            if not mb_id_str and not ben_id_str and not lieu and not date_str:
                continue

            # Ligne existante ?
            lien = None
            if mb_id_str:
                try:
                    mb_id_int = int(mb_id_str)
                    lien = MembreBenediction.query.get(mb_id_int)
                except ValueError:
                    lien = None

            # Si pas de bénédiction choisie ET qu'on a un lien existant => suppression
            if lien and not ben_id_str and not lieu and not date_str:
                db.session.delete(lien)
                continue

            # Si pas de bénédiction choisie mais des infos => on ignore (ou on pourrait flasher une erreur)
            if not ben_id_str:
                # Ici tu peux décider de faire un flash erreur si tu veux strict
                continue

            # Récupérer / vérifier la bénédiction
            try:
                ben_id_int = int(ben_id_str)
            except ValueError:
                continue

            ben = Benediction.query.get(ben_id_int)
            if not ben:
                continue

            # Créer le lien si inexistant
            if not lien:
                lien = MembreBenediction(
                    membre_id=membre.matricule,
                    benediction_id=ben.id
                )
                db.session.add(lien)

            # Mettre à jour les champs
            lien.benediction_id = ben.id
            lien.lieu = lieu or None
            if date_str:
                try:
                    lien.date_obtention = date.fromisoformat(date_str)
                except ValueError:
                    flash(f"Date de bénédiction invalide pour {ben.nom}.", "error")
            else:
                lien.date_obtention = None

        # =====================================================
        # 2) GESTION DES TALENTS (tableau mt_*)
        # =====================================================
        mt_ids = request.form.getlist('mt_id')
        mt_talent_ids = request.form.getlist('mt_talent_id')
        mt_comments = request.form.getlist('mt_commentaire')

        for idx in range(len(mt_ids)):
            mt_id_str = (mt_ids[idx] or "").strip()
            talent_id_str = (mt_talent_ids[idx] or "").strip()
            commentaire = (mt_comments[idx] or "").strip()

            # Ligne complètement vide => ignorer
            if not mt_id_str and not talent_id_str and not commentaire:
                continue

            lien_talent = None
            if mt_id_str:
                try:
                    mt_id_int = int(mt_id_str)
                    lien_talent = MembreTalent.query.get(mt_id_int)
                except ValueError:
                    lien_talent = None

            # Si aucun talent sélectionné et aucun commentaire, mais lien existant => suppression
            if lien_talent and not talent_id_str and not commentaire:
                db.session.delete(lien_talent)
                continue

            # Si pas de talent choisi mais commentaire : on peut ignorer (ou flasher une erreur)
            if not talent_id_str:
                continue

            try:
                talent_id_int = int(talent_id_str)
            except ValueError:
                continue

            talent_obj = Talent.query.get(talent_id_int)
            if not talent_obj:
                continue

            if not lien_talent:
                lien_talent = MembreTalent(
                    membre_id=membre.matricule,
                    talent_id=talent_obj.id
                )
                db.session.add(lien_talent)

            lien_talent.talent_id = talent_obj.id
            lien_talent.commentaire = commentaire or None

        # =====================================================
        # SAUVEGARDE
        # =====================================================
        db.session.commit()
        flash("Membre mis à jour.", "success")
        return redirect(url_for('membre_fiche', matricule=membre.matricule))

    return render_template(
        'membre_modifier.html',
        membre=membre,
        benedictions=benedictions,
        talents=talents,
        familles=familles
    )


    
@app.route('/membres/<matricule>/fiche')
def membre_fiche(matricule):
    membre = Membre.query.get_or_404(matricule)

    # Dernière inscription (pour afficher coordonnées "actuelles")
    derniere_inscription = None
    if membre.inscriptions:
        derniere_inscription = sorted(
            membre.inscriptions,
            key=lambda i: i.id,
            reverse=True
        )[0]

    # Historique des sessions où le membre est inscrit
    # On trie par date_debut décroissante (ou par id si pas de date)
    def session_sort_key(ins):
        s = ins.session
        if s and s.date_debut:
            return s.date_debut
        return date.min

    inscriptions_tries = sorted(
        membre.inscriptions,
        key=session_sort_key,
        reverse=True
    )

    # Historique des ministères + rôles par session
    ministeres_par_session = defaultdict(list)
    for smm in membre.ministeres:  # SessionMinistereMembre
        sm = smm.session_ministere
        if not sm:
            continue
        sess = sm.session
        if not sess:
            continue
        ministere_nom = sm.ministere.nom if sm.ministere else ''
        role_nom = smm.role.nom if smm.role else ''
        ministeres_par_session[sess.id].append({
            "ministere": ministere_nom,
            "role": role_nom,
        })

    return render_template(
        'membre_fiche.html',
        membre=membre,
        derniere_inscription=derniere_inscription,
        inscriptions_tries=inscriptions_tries,
        ministeres_par_session=ministeres_par_session
    )





# ---- CRUD Faritra / Benediction / Ministere / Talent / Role ---- #

@app.route('/faritra')
def faritra_liste():
    faritras = Faritra.query.order_by(Faritra.nom).all()
    return render_template('faritra_liste.html', faritras=faritras)


@app.route('/faritra/nouveau', methods=['GET', 'POST'])
def faritra_nouveau():
    if request.method == 'POST':
        nom = request.form.get('nom')
        if nom:
            f = Faritra(nom=nom)
            db.session.add(f)
            db.session.commit()
            flash("Faritra créé.", "success")
            return redirect(url_for('faritra_liste'))
    return render_template('faritra_form.html', faritra=None)


@app.route('/faritra/<int:faritra_id>/modifier', methods=['GET', 'POST'])
def faritra_modifier(faritra_id):
    faritra = Faritra.query.get_or_404(faritra_id)
    if request.method == 'POST':
        faritra.nom = request.form.get('nom')
        db.session.commit()
        flash("Faritra modifié.", "success")
        return redirect(url_for('faritra_liste'))
    return render_template('faritra_form.html', faritra=faritra)


@app.route('/faritra/<int:faritra_id>/supprimer', methods=['POST'])
def faritra_supprimer(faritra_id):
    faritra = Faritra.query.get_or_404(faritra_id)
    db.session.delete(faritra)
    db.session.commit()
    flash("Faritra supprimé.", "success")
    return redirect(url_for('faritra_liste'))


@app.route('/benedictions')
def benedictions_liste():
    benedictions = Benediction.query.order_by(Benediction.nom).all()
    return render_template('benedictions_liste.html', benedictions=benedictions)


@app.route('/benedictions/nouvelle', methods=['GET', 'POST'])
def benediction_nouvelle():
    if request.method == 'POST':
        nom = request.form.get('nom')
        if nom:
            b = Benediction(nom=nom)
            db.session.add(b)
            db.session.commit()
            flash("Bénédiction créée.", "success")
            return redirect(url_for('benedictions_liste'))
    return render_template('benediction_form.html', benediction=None)


@app.route('/benedictions/<int:benediction_id>/modifier', methods=['GET', 'POST'])
def benediction_modifier(benediction_id):
    benediction = Benediction.query.get_or_404(benediction_id)
    if request.method == 'POST':
        benediction.nom = request.form.get('nom')
        db.session.commit()
        flash("Bénédiction modifiée.", "success")
        return redirect(url_for('benedictions_liste'))
    return render_template('benediction_form.html', benediction=benediction)


@app.route('/benedictions/<int:benediction_id>/supprimer', methods=['POST'])
def benediction_supprimer(benediction_id):
    benediction = Benediction.query.get_or_404(benediction_id)
    db.session.delete(benediction)
    db.session.commit()
    flash("Bénédiction supprimée.", "success")
    return redirect(url_for('benedictions_liste'))


@app.route('/ministeres')
def ministeres_liste():
    ministeres = Ministere.query.order_by(Ministere.nom).all()
    return render_template('ministeres_liste.html', ministeres=ministeres)


@app.route('/ministeres/nouveau', methods=['GET', 'POST'])
def ministere_nouveau():
    if request.method == 'POST':
        nom = request.form.get('nom')
        if nom:
            m = Ministere(nom=nom)
            db.session.add(m)
            db.session.commit()
            flash("Ministère créé.", "success")
            return redirect(url_for('ministeres_liste'))
    return render_template('ministere_form.html', ministere=None)


@app.route('/ministeres/<int:ministere_id>/modifier', methods=['GET', 'POST'])
def ministere_modifier(ministere_id):
    ministere = Ministere.query.get_or_404(ministere_id)
    if request.method == 'POST':
        ministere.nom = request.form.get('nom')
        db.session.commit()
        flash("Ministère modifié.", "success")
        return redirect(url_for('ministeres_liste'))
    return render_template('ministere_form.html', ministere=ministere)


@app.route('/ministeres/<int:ministere_id>/supprimer', methods=['POST'])
def ministere_supprimer(ministere_id):
    ministere = Ministere.query.get_or_404(ministere_id)
    db.session.delete(ministere)
    db.session.commit()
    flash("Ministère supprimé.", "success")
    return redirect(url_for('ministeres_liste'))


@app.route('/talents')
def talents_liste():
    talents = Talent.query.order_by(Talent.nom).all()
    return render_template('talents_liste.html', talents=talents)


@app.route('/talents/nouveau', methods=['GET', 'POST'])
def talent_nouveau():
    if request.method == 'POST':
        nom = request.form.get('nom')
        if nom:
            t = Talent(nom=nom)
            db.session.add(t)
            db.session.commit()
            flash("Talent créé.", "success")
            return redirect(url_for('talents_liste'))
    return render_template('talent_form.html', talent=None)


@app.route('/talents/<int:talent_id>/modifier', methods=['GET', 'POST'])
def talent_modifier(talent_id):
    talent = Talent.query.get_or_404(talent_id)
    if request.method == 'POST':
        talent.nom = request.form.get('nom')
        db.session.commit()
        flash("Talent modifié.", "success")
        return redirect(url_for('talents_liste'))
    return render_template('talent_form.html', talent=talent)


@app.route('/talents/<int:talent_id>/supprimer', methods=['POST'])
def talent_supprimer(talent_id):
    talent = Talent.query.get_or_404(talent_id)
    db.session.delete(talent)
    db.session.commit()
    flash("Talent supprimé.", "success")
    return redirect(url_for('talents_liste'))


@app.route('/roles')
def roles_liste():
    roles = Role.query.order_by(Role.nom).all()
    return render_template('roles_liste.html', roles=roles)


@app.route('/roles/nouveau', methods=['GET', 'POST'])
def role_nouveau():
    if request.method == 'POST':
        nom = request.form.get('nom')
        if nom:
            r = Role(nom=nom)
            db.session.add(r)
            db.session.commit()
            flash("Rôle créé.", "success")
            return redirect(url_for('roles_liste'))
    return render_template('role_form.html', role=None)


@app.route('/roles/<int:role_id>/modifier', methods=['GET', 'POST'])
def role_modifier(role_id):
    role = Role.query.get_or_404(role_id)
    if request.method == 'POST':
        role.nom = request.form.get('nom')
        db.session.commit()
        flash("Rôle modifié.", "success")
        return redirect(url_for('roles_liste'))
    return render_template('role_form.html', role=role)


@app.route('/roles/<int:role_id>/supprimer', methods=['POST'])
def role_supprimer(role_id):
    role = Role.query.get_or_404(role_id)
    db.session.delete(role)
    db.session.commit()
    flash("Rôle supprimé.", "success")
    return redirect(url_for('roles_liste'))


# ---- Sessions ---- #

@app.route('/sessions')
def sessions_liste():
    page = request.args.get('page', 1, type=int)
    per_page = 20

    query = Session.query.order_by(Session.date_debut.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    sessions = pagination.items

    return render_template(
        'sessions_liste.html',
        sessions=sessions,
        pagination=pagination
    )

@app.route(
    '/sessions/<int:session_id>/ministeres/<int:session_ministere_id>/membres/<int:smm_id>/supprimer',
    methods=['POST']
)
def session_ministere_membre_supprimer(session_id, session_ministere_id, smm_id):
    session = Session.query.get_or_404(session_id)
    sm = SessionMinistere.query.get_or_404(session_ministere_id)

    # ⚠️ On eager-load le membre pour éviter tout lazy-load après
    smm = (
        SessionMinistereMembre.query
        .options(joinedload(SessionMinistereMembre.membre))
        .get_or_404(smm_id)
    )

    # sécurité : vérifier que le lien appartient bien à ce ministère + cette session
    if smm.session_ministere_id != sm.id or sm.session_id != session.id:
        flash("Incohérence de données : impossible de supprimer ce membre du ministère.", "error")
        return redirect(url_for('sessions_liste'))

    # On récupère le nom AVANT de supprimer (pour éviter DetachedInstanceError)
    nom_membre = smm.membre.nom if smm.membre else smm.membre_id

    db.session.delete(smm)
    db.session.commit()

    flash(f"Le membre {nom_membre} a été retiré de ce ministère.", "success")

    return redirect(url_for(
        'session_ministere_membres',
        session_id=session_id,
        session_ministere_id=session_ministere_id
    ))


@app.route('/sessions/nouvelle', methods=['GET', 'POST'])
def session_nouvelle():
    # On liste les sessions existantes pour la partie "copie"
    sessions_existantes = Session.query.order_by(Session.date_debut.desc()).all()

    # Pré-calcul des stats pour l’écran (facultatif, tu l’as déjà normalement)
    stats_sessions = {}
    for s in sessions_existantes:
        total_inscriptions = InscriptionSession.query.filter_by(session_id=s.id).count()
        total_ministeres = SessionMinistere.query.filter_by(session_id=s.id).count()
        total_membres_ministeres = (
            db.session.query(SessionMinistereMembre)
            .join(SessionMinistere, SessionMinistere.id == SessionMinistereMembre.session_ministere_id)
            .filter(SessionMinistere.session_id == s.id)
            .count()
        )
        stats_sessions[s.id] = {
            "inscriptions": total_inscriptions,
            "ministeres": total_ministeres,
            "membres_ministeres": total_membres_ministeres,
        }

    if request.method == 'POST':
        date_debut_str = request.form.get('date_debut')
        date_fin_str = request.form.get('date_fin')

        date_debut = parse_date(date_debut_str)
        date_fin = parse_date(date_fin_str)

        if not date_debut or not date_fin:
            flash("Dates de début et de fin obligatoires.", "error")
            return render_template('session_form.html',
                                   sessions_existantes=sessions_existantes,
                                   stats_sessions=stats_sessions)

        if date_fin < date_debut:
            flash("La date de fin doit être >= à la date de début.", "error")
            return render_template('session_form.html',
                                   sessions_existantes=sessions_existantes,
                                   stats_sessions=stats_sessions)

        # 1) Créer la nouvelle session (inactive au début)
        nouvelle_session = Session(
            date_debut=date_debut,
            date_fin=date_fin,
            active=False,
            est_cloture=False
        )
        db.session.add(nouvelle_session)
        db.session.flush()  # pour avoir nouvelle_session.id

        # Récupérer les choix de l'utilisateur
        source_session_id_str = request.form.get('source_session_id', '').strip()
        copier_inscriptions = 'copier_inscriptions' in request.form
        copier_ministeres = 'copier_ministeres' in request.form

        previous_session = None
        if source_session_id_str:
            try:
                source_session_id = int(source_session_id_str)
                previous_session = Session.query.get(source_session_id)
            except ValueError:
                previous_session = None

        nb_membres_copies = 0
        nb_ministeres_copies = 0
        nb_membres_ministeres_copies = 0

        # 2) Copie des inscriptions
        if previous_session and copier_inscriptions:
            inscriptions_prev = InscriptionSession.query.filter_by(
                session_id=previous_session.id
            ).all()

            for ins in inscriptions_prev:
                m = ins.membre
                if not m:
                    continue
                if m.est_mort or m.date_depart is not None:
                    # ne pas reporter les membres morts ou partis
                    continue

                new_ins = InscriptionSession(
                    session_id=nouvelle_session.id,
                    membre_id=m.matricule,
                    adresse=ins.adresse,
                    faritra_id=ins.faritra_id,
                    telephone=ins.telephone,
                    responsable_id=ins.responsable_id
                )
                db.session.add(new_ins)
                nb_membres_copies += 1

        # 3) Copie des ministères + rôles + membres de ministères
        if previous_session and copier_ministeres:
            # Rôle par défaut "Membre" (optionnel pour les membres de ministères)
            role_default = (
                Role.query
                .filter(Role.nom.ilike("membre"))
                .first()
            )

            # 3.a. Copier les SessionMinistere (ministères de la session)
            sm_mapping = {}  # old_sm.id -> new_sm
            session_ministeres_prev = SessionMinistere.query.filter_by(
                session_id=previous_session.id
            ).all()

            for sm_prev in session_ministeres_prev:
                new_sm = SessionMinistere(
                    session_id=nouvelle_session.id,
                    ministere_id=sm_prev.ministere_id
                )
                db.session.add(new_sm)
                db.session.flush()
                sm_mapping[sm_prev.id] = new_sm
                nb_ministeres_copies += 1

            # 3.b. Copier la configuration des rôles par ministère (SessionMinistereRole)
            old_sm_ids = list(sm_mapping.keys())
            if old_sm_ids:
                smr_prev_all = SessionMinistereRole.query.filter(
                    SessionMinistereRole.session_ministere_id.in_(old_sm_ids)
                ).all()

                for smr_prev in smr_prev_all:
                    new_sm = sm_mapping.get(smr_prev.session_ministere_id)
                    if not new_sm:
                        continue

                    smr_new = SessionMinistereRole(
                        session_ministere_id=new_sm.id,
                        role_id=smr_prev.role_id
                    )
                    db.session.add(smr_new)

            # 3.c. Copier les membres + rôles affectés aux ministères (SessionMinistereMembre)
            smm_prev_all = (
                SessionMinistereMembre.query
                .join(SessionMinistere, SessionMinistere.id == SessionMinistereMembre.session_ministere_id)
                .filter(SessionMinistere.session_id == previous_session.id)
                .all()
            )

            for smm_prev in smm_prev_all:
                m = smm_prev.membre
                if not m:
                    continue
                if m.est_mort or m.date_depart is not None:
                    continue

                old_sm_id = smm_prev.session_ministere_id
                new_sm = sm_mapping.get(old_sm_id)
                if not new_sm:
                    continue

                # Copier le rôle existant, ou mettre le rôle "Membre" par défaut si aucun
                new_role_id = smm_prev.role_id
                if not new_role_id and role_default:
                    new_role_id = role_default.id

                smm_new = SessionMinistereMembre(
                    session_ministere_id=new_sm.id,
                    membre_id=m.matricule,
                    role_id=new_role_id
                )
                db.session.add(smm_new)
                nb_membres_ministeres_copies += 1

        # 4) Activer la nouvelle session et désactiver les autres
        Session.query.update({Session.active: False})
        nouvelle_session.active = True

        db.session.commit()

        # 5) Message récap
        if previous_session:
            msg = f"Session créée et activée (copie depuis session {previous_session.id} : "
            if copier_inscriptions:
                msg += f"{nb_membres_copies} inscriptions copiées, "
            if copier_ministeres:
                msg += f"{nb_ministeres_copies} ministères, {nb_membres_ministeres_copies} membres de ministères copiés"
            msg = msg.rstrip(", ") + ")."
            flash(msg, "success")
        else:
            flash("Session créée et activée (aucune session source sélectionnée).", "success")

        return redirect(url_for('sessions_liste'))

    # GET : afficher le formulaire avec la liste des sessions comme source possible
    return render_template(
        'session_form.html',
        sessions_existantes=sessions_existantes,
        stats_sessions=stats_sessions
    )


@app.route('/sessions/<int:session_id>/modifier', methods=['GET', 'POST'])
def session_modifier(session_id):
    s = Session.query.get_or_404(session_id)

    if request.method == 'POST':
        s.date_debut = parse_date(request.form.get('date_debut'))
        s.date_fin = parse_date(request.form.get('date_fin'))
        s.active = 'active' in request.form
        s.est_cloture = 'est_cloture' in request.form

        db.session.commit()
        flash("Session mise à jour.", "success")
        return redirect(url_for('sessions_liste'))

    return render_template('sessions_form.html', session=s)


@app.route('/sessions/<int:session_id>/supprimer', methods=['POST'])
def session_supprimer(session_id):
    session_obj = Session.query.get_or_404(session_id)

    # On garde en mémoire si elle était active
    etait_active = session_obj.active

    # 1) Récupérer les SessionMinistere de cette session
    sm_list = SessionMinistere.query.filter_by(session_id=session_id).all()
    sm_ids = [sm.id for sm in sm_list]

    # 2) Supprimer les liens membres ↔ ministères de cette session
    if sm_ids:
        SessionMinistereMembre.query.filter(
            SessionMinistereMembre.session_ministere_id.in_(sm_ids)
        ).delete(synchronize_session=False)

        # 3) Supprimer les ministères de cette session
        SessionMinistere.query.filter(
            SessionMinistere.id.in_(sm_ids)
        ).delete(synchronize_session=False)

    # 4) Supprimer les inscriptions de cette session
    InscriptionSession.query.filter_by(session_id=session_id).delete(synchronize_session=False)

    # 5) Supprimer la session elle-même
    db.session.delete(session_obj)
    db.session.commit()

    # 6) Si c'était la session active, on en active une autre (la plus récente non clôturée si possible)
    if etait_active:
        nouvelle_active = (
            Session.query
            .filter_by(est_cloture=False)
            .order_by(Session.date_debut.desc())
            .first()
        )
        if nouvelle_active:
            nouvelle_active.active = True
            db.session.commit()

    flash(f"La session {session_id} et toutes ses données associées ont été supprimées.", "success")
    return redirect(url_for('sessions_liste'))



# ---- Ministères par session ---- #

@app.route('/sessions/<int:session_id>/ministeres', methods=['GET', 'POST'])
def session_ministeres(session_id):
    session = Session.query.get_or_404(session_id)
    session_ministeres = SessionMinistere.query.filter_by(session_id=session_id).all()
    tous_ministeres = Ministere.query.order_by(Ministere.nom).all()

    if request.method == 'POST':
        ministere_id = request.form.get('ministere_id')
        if ministere_id:
            sm = SessionMinistere(session_id=session_id, ministere_id=ministere_id)
            db.session.add(sm)
            db.session.commit()
            flash("Ministère ajouté à la session.", "success")
            return redirect(url_for('session_ministeres', session_id=session_id))

    return render_template(
        'session_ministeres.html',
        session=session,
        session_ministeres=session_ministeres,
        tous_ministeres=tous_ministeres
    )

@app.route('/sessions/<int:session_id>/ministeres/<int:session_ministere_id>/roles',
           methods=['GET', 'POST'])
def session_ministere_roles(session_id, session_ministere_id):
    session = Session.query.get_or_404(session_id)
    session_ministere = SessionMinistere.query.get_or_404(session_ministere_id)

    # Sécurité : vérifier cohérence session <-> ministère
    if session_ministere.session_id != session.id:
        flash("Ce ministère n'appartient pas à cette session.", "error")
        return redirect(url_for('sessions_liste'))

    # Rôles déjà associés à ce ministère (session)
    roles_associes = (
	    SessionMinistereRole.query
	    .join(Role, Role.id == SessionMinistereRole.role_id)
	    .filter(SessionMinistereRole.session_ministere_id == session_ministere_id)
	    .order_by(SessionMinistereRole.ordre.asc().nulls_last(), Role.nom.asc())
	    .all()
	)


    # Liste des rôles disponibles qui ne sont pas encore associés
    roles_associes_ids = [smr.role_id for smr in roles_associes]
    if roles_associes_ids:
        roles_disponibles = (
            Role.query
            .filter(~Role.id.in_(roles_associes_ids))
            .order_by(Role.nom)
            .all()
        )
    else:
        roles_disponibles = Role.query.order_by(Role.nom).all()

    # Gestion du POST
    if request.method == 'POST':
        action = request.form.get('action', '').strip()

        # ----------------------------------------------
        # 1) Ajouter un rôle
        # ------------------------------
        if action == 'add':
            new_role_id = request.form.get('new_role_id', '').strip()
            if not new_role_id:
                flash("Veuillez choisir un rôle à ajouter.", "error")
                return redirect(url_for('session_ministere_roles',
                                        session_id=session_id,
                                        session_ministere_id=session_ministere_id))

            try:
                new_role_id_int = int(new_role_id)
            except:
                flash("Rôle invalide.", "error")
                return redirect(url_for('session_ministere_roles',
                                        session_id=session_id,
                                        session_ministere_id=session_ministere_id))

            # Vérifier si déjà présent
            deja = SessionMinistereRole.query.filter_by(
                session_ministere_id=session_ministere_id,
                role_id=new_role_id_int
            ).first()

            if deja:
                flash("Ce rôle est déjà associé à ce ministère.", "warning")
            else:
                smr_new = SessionMinistereRole(
                    session_ministere_id=session_ministere_id,
                    role_id=new_role_id_int,
                    max_membres=None,
                    ordre=None
                )
                db.session.add(smr_new)
                db.session.commit()
                flash("Rôle ajouté à ce ministère pour cette session.", "success")

            return redirect(url_for('session_ministere_roles',
                                    session_id=session_id,
                                    session_ministere_id=session_ministere_id))

        # ------------------------------
        # 2) Supprimer un ou plusieurs rôles
        # ------------------------------
        if action == 'delete':
            to_delete = request.form.getlist('supprimer_role_ids')
            if to_delete:
                ids_int = [int(x) for x in to_delete]

                SessionMinistereRole.query.filter(
                    SessionMinistereRole.id.in_(ids_int),
                    SessionMinistereRole.session_ministere_id == session_ministere_id
                ).delete(synchronize_session=False)

                db.session.commit()
                flash(f"{len(ids_int)} rôle(s) supprimé(s).", "success")
            else:
                flash("Aucun rôle sélectionné pour suppression.", "warning")

            return redirect(url_for('session_ministere_roles',
                                    session_id=session_id,
                                    session_ministere_id=session_ministere_id))

        # ------------------------------
        # 3) Mettre à jour max_membres + ordre
        # ------------------------------
        if action == 'update_settings':
            for smr in roles_associes:

                # --- max_membres ---
                max_field = f"max_{smr.id}"
                max_val = request.form.get(max_field, '').strip()

                if max_val == "":
                    smr.max_membres = None
                else:
                    try:
                        smr.max_membres = int(max_val)
                    except:
                        pass

                # --- ordre ---
                ordre_field = f"ordre_{smr.id}"
                ordre_val = request.form.get(ordre_field, '').strip()

                if ordre_val == "":
                    smr.ordre = None
                else:
                    try:
                        smr.ordre = int(ordre_val)
                    except:
                        pass

            db.session.commit()
            flash("Paramètres des rôles mis à jour.", "success")

            return redirect(url_for('session_ministere_roles',
                                    session_id=session_id,
                                    session_ministere_id=session_ministere_id))

    # ------------------------------
    # Affichage GET
    # ------------------------------
    return render_template(
        'session_ministere_roles.html',
        session=session,
        session_ministere=session_ministere,
        roles_associes=roles_associes,
        roles_disponibles=roles_disponibles
    )


@app.route('/sessions/<int:session_id>/ministeres/<int:session_ministere_id>/membres', methods=['GET', 'POST'])
def session_ministere_membres(session_id, session_ministere_id):
    session = Session.query.get_or_404(session_id)
    sm = SessionMinistere.query.get_or_404(session_ministere_id)

    if sm.session_id != session.id:
        flash("Ce ministère n'appartient pas à cette session.", "error")
        return redirect(url_for('sessions_liste'))

    # ⚠️ On charge a.membre et a.role en eager-load pour éviter le DetachedInstanceError
    affectations = (
	    SessionMinistereMembre.query
	    .join(SessionMinistereRole, SessionMinistereRole.role_id == SessionMinistereMembre.role_id)
	    .join(Membre, Membre.matricule == SessionMinistereMembre.membre_id)
	    .options(
		joinedload(SessionMinistereMembre.membre),
		joinedload(SessionMinistereMembre.role)
	    )
	    .filter(SessionMinistereMembre.session_ministere_id == sm.id)
	    .order_by(SessionMinistereRole.ordre.asc().nulls_last(), Membre.nom.asc())
	    .all()
	)


    # Membres éligibles à l'affectation (à adapter selon ta logique métier si besoin)
    # IDs des membres déjà dans CE ministère (pour ne pas les proposer deux fois)
    deja_membre_ids = [a.membre_id for a in affectations]

    # On cherche le ministère "FDL"
    ministere_fdl = Ministere.query.filter(Ministere.nom.ilike("FDL")).first()
    membres = []

    if ministere_fdl:
        # SessionMinistere correspondant à FDL pour cette session
        sm_fdl = SessionMinistere.query.filter_by(
            session_id=session.id,
            ministere_id=ministere_fdl.id
        ).first()

        if sm_fdl:
            # Tous les membres affectés au ministère FDL dans cette session
            smm_fdl = SessionMinistereMembre.query.filter_by(
                session_ministere_id=sm_fdl.id
            ).all()
            fdl_membre_ids = [smm.membre_id for smm in smm_fdl]

            if fdl_membre_ids:
                # On ne garde que :
                # - membres FDL
                # - non morts
                # - pas déjà dans ce ministère
                membres = (
                    Membre.query
                    .filter(
                        Membre.matricule.in_(fdl_membre_ids),
                        Membre.est_mort == False,
                        ~Membre.matricule.in_(deja_membre_ids)
                    )
                    .order_by(Membre.nom)
                    .all()
                )

    # Si pas de ministère FDL ou pas de SessionMinistere FDL : fallback (optionnel)
    if not membres:
        membres = (
            Membre.query
            .join(InscriptionSession, InscriptionSession.membre_id == Membre.matricule)
            .filter(
                InscriptionSession.session_id == session.id,
                Membre.est_mort == False,
                ~Membre.matricule.in_(deja_membre_ids)
            )
            .order_by(Membre.nom)
            .all()
        )

    # Rôles disponibles pour ce ministère (configurés via SessionMinistereRole)
    roles_disponibles = (
        Role.query
        .join(SessionMinistereRole, SessionMinistereRole.role_id == Role.id)
        .filter(SessionMinistereRole.session_ministere_id == sm.id)
        .order_by(Role.nom)
        .all()
    )
    
    # config des limites par rôle pour ce ministère
    roles_config = (
	    SessionMinistereRole.query
	    .join(Role, Role.id == SessionMinistereRole.role_id)
	    .filter(SessionMinistereRole.session_ministere_id == sm.id)
	    .order_by(SessionMinistereRole.ordre.asc().nulls_last(), Role.nom.asc())
	    .all()
	)

    limits_by_role = {rc.role_id: rc.max_membres for rc in roles_config}
    
    # Récap par rôle : nombre de membres actuels / max
    role_stats = []
    for rc in roles_config:
        current_count = SessionMinistereMembre.query.filter_by(
            session_ministere_id=sm.id,
            role_id=rc.role_id
        ).count()
        role_stats.append({
            "role_nom": rc.role.nom,
            "count": current_count,
            "max": rc.max_membres  # peut être None (illimité)
        })



    if request.method == 'POST':
        action = request.form.get('action', '').strip()

        # 1) Mise à jour des rôles des membres déjà affectés
        if action == 'update_roles':
            for a in affectations:
                field_name = f"role_id_{a.id}"
                role_id_str = request.form.get(field_name, '').strip()

                new_role_id = int(role_id_str) if role_id_str else None
                old_role_id = a.role_id

                # Si le rôle ne change pas -> pas besoin de vérifier
                if new_role_id == old_role_id:
                    continue

                # Si le nouveau rôle est limité, vérifier le quota
                if new_role_id is not None:
                    max_m = limits_by_role.get(new_role_id)
                    if max_m is not None:
                        current_count = SessionMinistereMembre.query.filter_by(
                            session_ministere_id=sm.id,
                            role_id=new_role_id
                        ).count()

                        # a n'a pas encore ce rôle (puisqu'on est dans le cas new != old),
                        # donc si current_count >= max_m, on refuse.
                        if current_count >= max_m:
                            flash(
                                f"Impossible d'assigner le rôle à {a.membre.nom} : "
                                f"quota maximum déjà atteint.",
                                "error"
                            )
                            return redirect(url_for('session_ministere_membres',
                                                    session_id=session_id,
                                                    session_ministere_id=session_ministere_id))

                # OK, on peut appliquer le changement
                a.role_id = new_role_id

            db.session.commit()
            flash("Rôles des membres du ministère mis à jour.", "success")
            return redirect(url_for('session_ministere_membres',
                                    session_id=session_id,
                                    session_ministere_id=session_ministere_id))

        # 2) Ajout d'un nouveau membre au ministère
        if action == 'ajout':
            membre_id = request.form.get('membre_id', '').strip()
            role_id_str = request.form.get('role_id', '').strip()

            if not membre_id:
                flash("Veuillez choisir un membre à affecter.", "error")
                return redirect(url_for('session_ministere_membres',
                                        session_id=session_id,
                                        session_ministere_id=session_ministere_id))

            # Vérifier s'il est déjà affecté à ce ministère
            deja = SessionMinistereMembre.query.filter_by(
                session_ministere_id=sm.id,
                membre_id=membre_id
            ).first()
            if deja:
                flash("Ce membre est déjà affecté à ce ministère.", "warning")
                return redirect(url_for('session_ministere_membres',
                                        session_id=session_id,
                                        session_ministere_id=session_ministere_id))

            role_id = int(role_id_str) if role_id_str else None

            # Vérifier quota si rôle sélectionné et limité
            if role_id is not None:
                max_m = limits_by_role.get(role_id)
                if max_m is not None:
                    current_count = SessionMinistereMembre.query.filter_by(
                        session_ministere_id=sm.id,
                        role_id=role_id
                    ).count()
                    if current_count >= max_m:
                        flash("Le nombre maximum de membres pour ce rôle est déjà atteint.", "error")
                        return redirect(url_for('session_ministere_membres',
                                                session_id=session_id,
                                                session_ministere_id=session_ministere_id))


            smm_new = SessionMinistereMembre(
                session_ministere_id=sm.id,
                membre_id=membre_id,
                role_id=role_id
            )
            db.session.add(smm_new)
            db.session.commit()
            flash("Membre affecté au ministère.", "success")

            return redirect(url_for('session_ministere_membres',
                                    session_id=session_id,
                                    session_ministere_id=session_ministere_id))

    return render_template(
        'session_ministere_membres.html',
        session=session,
        sm=sm,
        affectations=affectations,
        membres=membres,
        roles_disponibles=roles_disponibles,
        role_stats=role_stats
    )


# ---- Inscriptions ---- #

def get_responsables_fdl(session_id):
    return db.session.query(Membre).join(
        SessionMinistereMembre, SessionMinistereMembre.membre_id == Membre.matricule
    ).join(
        SessionMinistere, SessionMinistere.id == SessionMinistereMembre.session_ministere_id
    ).join(
        Ministere, Ministere.id == SessionMinistere.ministere_id
    ).filter(
        SessionMinistere.session_id == session_id,
        Ministere.nom == 'FDL'
    ).distinct().order_by(Membre.nom).all()


@app.route('/sessions/<int:session_id>/inscriptions')
def inscriptions_liste(session_id):
    session = Session.query.get_or_404(session_id)
    inscriptions = InscriptionSession.query.filter_by(session_id=session_id).all()

    # Membres éligibles = non morts
    membres_total = Membre.query.filter_by(est_mort=False).all()

    # Combien de membres (non morts) déjà inscrits ?
    inscrits_ids = {
        ins.membre_id
        for ins in inscriptions
        if ins.membre and not ins.membre.est_mort
    }
    nb_inscrits = len(inscrits_ids)
    nb_possibles = len(membres_total)
    nb_restants = max(nb_possibles - nb_inscrits, 0)

    return render_template(
        'inscriptions_liste.html',
        session=session,
        inscriptions=inscriptions,
        membres_total=membres_total,
        nb_inscrits=nb_inscrits,
        nb_possibles=nb_possibles,
        nb_restants=nb_restants
    )





@app.route('/sessions/<int:session_id>/inscriptions/nouvelle', methods=['GET', 'POST'])
def inscription_nouvelle(session_id):
    session = Session.query.get_or_404(session_id)

    # Membres éligibles : non inscrits à cette session, non morts
    subq = db.session.query(InscriptionSession.membre_id).filter_by(session_id=session_id)
    membres = Membre.query.filter(
        ~Membre.matricule.in_(subq),
        Membre.est_mort == False
    ).order_by(Membre.nom).all()

    faritras = Faritra.query.order_by(Faritra.nom).all()
    responsables = get_responsables_fdl(session_id)

    if request.method == 'POST':
        membre_id = request.form.get('membre_id')
        adresse = request.form.get('adresse', '').strip()
        faritra_id = request.form.get('faritra_id') or None
        telephone = request.form.get('telephone', '').strip()
        responsable_id = request.form.get('responsable_id') or None

        errors = []
        if not membre_id:
            errors.append("Le membre est obligatoire.")
        if not adresse:
            errors.append("L'adresse est obligatoire.")
        if not faritra_id:
            errors.append("Le faritra est obligatoire.")
        if not telephone:
            errors.append("Le téléphone est obligatoire.")

        if errors:
            for e in errors:
                flash(e, "error")
            # On ré-affiche le formulaire avec les erreurs
            return render_template(
                'inscriptions_form.html',
                session=session,
                inscription=None,
                membres=membres,
                faritras=faritras,
                responsables=responsables
            )

        insc = InscriptionSession(
            session_id=session_id,
            membre_id=membre_id,
            adresse=adresse,
            faritra_id=faritra_id,
            telephone=telephone,
            responsable_id=responsable_id
        )
        db.session.add(insc)
        db.session.commit()
        flash("Inscription ajoutée.", "success")
        return redirect(url_for('inscriptions_liste', session_id=session_id))

    return render_template(
        'inscriptions_form.html',
        session=session,
        inscription=None,
        membres=membres,
        faritras=faritras,
        responsables=responsables
    )



@app.route('/sessions/<int:session_id>/inscriptions/<int:inscription_id>/modifier',
           methods=['GET', 'POST'])
def inscription_modifier(session_id, inscription_id):
    session = Session.query.get_or_404(session_id)
    inscription = InscriptionSession.query.get_or_404(inscription_id)

    # Membres éligibles :
    # - pas déjà inscrits à cette session (sauf ce même enregistrement)
    # - non morts
    subq = db.session.query(InscriptionSession.membre_id).filter(
        InscriptionSession.session_id == session_id,
        InscriptionSession.id != inscription_id
    )

    membres = Membre.query.filter(
        (
            (Membre.matricule == inscription.membre_id) |
            (~Membre.matricule.in_(subq))
        ),
        Membre.est_mort == False
    ).order_by(Membre.nom).all()

    faritras = Faritra.query.order_by(Faritra.nom).all()
    responsables = get_responsables_fdl(session_id)

    if request.method == 'POST':
        membre_id = request.form.get('membre_id')
        adresse = request.form.get('adresse', '').strip()
        faritra_id = request.form.get('faritra_id') or None
        telephone = request.form.get('telephone', '').strip()
        responsable_id = request.form.get('responsable_id') or None

        errors = []
        if not membre_id:
            errors.append("Le membre est obligatoire.")
        if not adresse:
            errors.append("L'adresse est obligatoire.")
        if not faritra_id:
            errors.append("Le faritra est obligatoire.")
        if not telephone:
            errors.append("Le téléphone est obligatoire.")

        if errors:
            for e in errors:
                flash(e, "error")
            return render_template(
                'inscriptions_form.html',
                session=session,
                inscription=inscription,
                membres=membres,
                faritras=faritras,
                responsables=responsables
            )

        inscription.membre_id = membre_id
        inscription.adresse = adresse
        inscription.faritra_id = faritra_id
        inscription.telephone = telephone
        inscription.responsable_id = responsable_id

        db.session.commit()
        flash("Inscription modifiée.", "success")
        return redirect(url_for('inscriptions_liste', session_id=session_id))

    return render_template(
        'inscriptions_form.html',
        session=session,
        inscription=inscription,
        membres=membres,
        faritras=faritras,
        responsables=responsables
    )



@app.route('/sessions/<int:session_id>/inscriptions/<int:inscription_id>/supprimer',
           methods=['POST'])
def inscription_supprimer(session_id, inscription_id):
    inscription = InscriptionSession.query.get_or_404(inscription_id)
    db.session.delete(inscription)
    db.session.commit()
    flash("Inscription supprimée.", "success")
    return redirect(url_for('inscriptions_liste', session_id=session_id))


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        seed_data()
    app.run(debug=True)
